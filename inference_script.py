#!/usr/bin/env python3
"""
MolmoAct2 - Polymetis Bridge (Joint Position)
Connects MolmoAct2 policy server to Franka robot via Polymetis (NUC)

Outputs ABSOLUTE joint positions, action_dim=8: [q1..q7, gripper]
  - q1-q7: Absolute joint positions in radians
  - gripper: forwarded raw as the robot gripper command (no clip or remap in this bridge)

Usage:
  1. Set MOLMOACT2_URL and PROMPT in Configuration section below
  2. Start the MolmoAct2 server (see molmoact2_server.py)
  3. Run this bridge: python molmoact2_real.py
"""
import time
import sys
import os
import signal
import atexit
import numpy as np
import cv2
import requests
import json_numpy
from datetime import datetime
from pathlib import Path

json_numpy.patch()

# Excel logging imports (only used if ENABLE_EXCEL_LOGGING is True)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Global flag for emergency stop
_emergency_stop = False

def signal_handler(signum, frame):
    """Handle Ctrl+C by setting emergency stop flag"""
    global _emergency_stop
    _emergency_stop = True
    print("\n\n[STOP] Ctrl+C detected - Emergency stop triggered!", flush=True)
    raise KeyboardInterrupt  # Raise to interrupt input() calls

# Register signal handler
signal.signal(signal.SIGINT, signal_handler)

# Add project path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

# Import ZED SDK
try:
    import pyzed.sl as sl
    ZED_AVAILABLE = True
except ImportError:
    print("Warning: ZED SDK not found. ZED camera will not be available.")
    ZED_AVAILABLE = False

# ============ Configuration ============
NUC_IP = "192.168.1.6"              # Polymetis zerorpc server IP (port 4242)
MOLMOACT2_URL = "http://unachievable-tawana-subtransparent.ngrok-free.dev"  # MolmoAct2 policy server URL
PROMPT = "pick up the blue block."

# ============ Loop Recording Configuration ============
# Set LOOP to the number of episodes you want to record
# LOOP = 0: Stop record (inference only, no video)
# LOOP = 1: Record 1 episode (single recording)
# LOOP = 3: Record 3 episodes (will loop 3 times)
# Each episode will be saved in a separate timestamped folder under a task-named directory
LOOP = 3  # Number of episodes to record per position per model

# ============ Position Configuration ============
# New recording scheme:
# - 6 models (3 velocity: pi0, pi0_fast, pi05 + 3 position: pi0, pi0_fast, pi05)
# - Each model runs 9 times: 3 initial positions × 3 videos per position
# - Total: 6 models × 9 runs = 54 videos per task
ENABLE_POSITION_VARIANT = True  # Set to True to use position variants and log to Excel
POSITION_VARIANT = "pos-1"  # Current position: "pos-1", "pos-2", or "pos-3" (only used if ENABLE_POSITION_VARIANT is True)

# ============ Excel Logging Configuration ============
# Set to True to enable Excel logging of episode metadata
# When enabled, you will be prompted after each episode for:
# - Model name (e.g., "pi0", "pi05") - asked once at start
# - Success status (y/n) - asked after each episode
# Data logged: Task, Episode, Model, Success, Video Path, Steps
# Excel file will be saved in the vid directory as "episode_log.xlsx"
ENABLE_EXCEL_LOGGING = True  # True: Enable logging, False: Disable

# ZED Camera Configuration (Official OpenPI DROID Setup)
# Both cameras now working on USB 3.0
# Using Serial Numbers instead of IDs for stability (IDs can change after reconnection)
# If a camera was replaced, re-check current serial numbers with:
#   python - <<'PY'
#   import pyzed.sl as sl
#   for d in sl.Camera.get_device_list():
#       print(d.camera_model, d.serial_number, d.camera_state, d.path)
#   PY
# If a camera shows up as NOT AVAILABLE, inspect USB-level serials with:
#   for dev in $(lsusb -d 2b03: | awk '{print $6}'); do
#       echo "== $dev =="; lsusb -v -d "$dev" 2>/dev/null | rg 'iProduct|iSerial'
#   done
ZED_EXTERNAL_ID = None              # Not using ID, using SN instead
ZED_EXTERNAL_SN = 26706125          # ZED 2 (external/shoulder view)
ZED_WRIST_ID = None                 # Not using ID, using SN instead
ZED_WRIST_SN = 15679333             # ZED Mini (wrist view)
WIDTH, HEIGHT, FPS = 1280, 720, 15  # ZED HD720 mode @ 15fps (official DROID config)
# Note: Using SN ensures cameras are always correctly identified even after USB reconnection

# Control parameters
CTRL_HZ = 5.0                      # Control frequency (Hz) - balanced speed and smoothness
# Note: Video recording uses camera FPS (15fps), not control frequency

# MolmoAct2 outputs ABSOLUTE joint positions, action_dim=8 ([q1..q7, gripper])
OPEN_LOOP_HORIZON = 15           # Use first 8 actions from the chunk
MAX_DQ = 0.15                   # Max joint delta per step (rad) for safety clipping
OUTPUT_IS_DELTA = False         # MolmoAct2 outputs absolute positions
USE_EMA_SMOOTHING = True        # Smooth absolute positions to reduce jumps
DELTA_SCALE = 1.0               # No scaling needed for absolute positions
EMA_ALPHA = 0.7                 # EMA smoothing factor (0.7 = faster response, less lag)

# Input image size expected by MolmoAct2 server
INPUT_WIDTH = 320
INPUT_HEIGHT = 180

# Maximum steps before stopping each episode
MAX_STEPS = 1000

print(f"Policy: MolmoAct2 (server: {MOLMOACT2_URL})")
print(f"Control frequency: {CTRL_HZ} Hz")
print(f"Open loop horizon: {OPEN_LOOP_HORIZON} actions")
print(f"Max joint delta: {MAX_DQ} rad")
print(f"Max steps per episode: {MAX_STEPS}")
print(f"Output type: ABSOLUTE (positions)")
print(f"EMA smoothing: {'enabled (alpha=' + str(EMA_ALPHA) + ')' if USE_EMA_SMOOTHING else 'disabled'}")
# =======================================

def print_zed_sn_lookup_help():
    """Print commands for checking current ZED serial numbers."""
    print("  To check current ZED serial numbers:")
    print("    python - <<'PY'")
    print("    import pyzed.sl as sl")
    print("    for d in sl.Camera.get_device_list():")
    print("        print(d.camera_model, d.serial_number, d.camera_state, d.path)")
    print("    PY")
    print("  If a camera shows NOT AVAILABLE, check USB-level serials:")
    print("    for dev in $(lsusb -d 2b03: | awk '{print $6}'); do")
    print("        echo \"== $dev ==\"; lsusb -v -d \"$dev\" 2>/dev/null | rg 'iProduct|iSerial'")
    print("    done")

def print_zed_device_list():
    """Print the current device list seen by the ZED SDK."""
    if not ZED_AVAILABLE:
        return
    try:
        devices = sl.Camera.get_device_list()
    except Exception as exc:
        print(f"  Could not query ZED device list: {exc}")
        return
    print("  ZED SDK device list:")
    if not devices:
        print("    <empty>")
        return
    for i, dev in enumerate(devices):
        print(f"    [{i}] model={dev.camera_model} sn={dev.serial_number} state={dev.camera_state} path={dev.path}")

# Global camera references for cleanup
_zed_ext = None
_zed_wri = None

def cleanup_cameras():
    """Cleanup function to ensure cameras are properly closed"""
    global _zed_ext, _zed_wri
    
    print("\nCleaning up cameras...")
    try:
        if _zed_ext is not None:
            _zed_ext.close()
            print("  [OK] External ZED camera closed")
            _zed_ext = None
    except Exception as e:
        print(f"  [WARN] Error closing external camera: {e}")
    
    try:
        if _zed_wri is not None:
            _zed_wri.close()
            print("  [OK] Wrist ZED camera closed")
            _zed_wri = None
    except Exception as e:
        print(f"  [WARN] Error closing wrist camera: {e}")
    
    # Add a small delay to ensure cleanup completes
    time.sleep(0.5)

def signal_handler(sig, frame):
    """Handle termination signals"""
    print(f"\n\nReceived signal {sig}")
    cleanup_cameras()
    sys.exit(0)

# Register cleanup handlers
atexit.register(cleanup_cameras)
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # kill command


def open_zed_camera(camera_id=None, serial_number=None, width=1280, height=720, fps=15):
    """Open ZED camera with optimized settings
    
    Args:
        camera_id: Camera ID (0 for first camera, 1 for second, etc.)
        serial_number: Camera serial number (alternative to camera_id)
        width, height: Resolution (1280x720 for HD720)
        fps: Frame rate (15 fps for DROID)
    """
    if not ZED_AVAILABLE:
        raise RuntimeError("ZED SDK is not available. Please install pyzed.")
    
    if camera_id is None and serial_number is None:
        raise ValueError("Must provide either camera_id or serial_number")
    
    zed = sl.Camera()
    
    # Set initialization parameters
    init_params = sl.InitParameters()
    init_params.camera_resolution = sl.RESOLUTION.HD720  # 1280x720
    init_params.camera_fps = fps
    init_params.depth_mode = sl.DEPTH_MODE.NONE  # We only need RGB, no depth
    init_params.coordinate_units = sl.UNIT.METER
    
    # Set camera ID or serial number
    if camera_id is not None:
        init_params.set_from_camera_id(camera_id)
        identifier = f"ID {camera_id}"
    else:
        init_params.set_from_serial_number(serial_number)
        identifier = f"SN {serial_number}"
    
    # Open the camera
    err = zed.open(init_params)
    if err != sl.ERROR_CODE.SUCCESS:
        raise RuntimeError(f"Failed to open ZED camera {identifier}: {err}")
    
    # Get camera information
    cam_info = zed.get_camera_information()
    camera_model = cam_info.camera_model
    actual_serial_number = cam_info.serial_number
    
    # Set camera controls for better image quality
    zed.set_camera_settings(sl.VIDEO_SETTINGS.WHITEBALANCE_AUTO, 1)  # Enable auto white balance
    
    # Set runtime parameters
    runtime_params = sl.RuntimeParameters()
    
    return zed, runtime_params, camera_model, actual_serial_number


def get_zed_image(zed, runtime_params, rotate_180=False):
    """Get RGB image from ZED camera
    
    Args:
        zed: ZED camera object
        runtime_params: Runtime parameters
        rotate_180: Whether to rotate image 180 degrees (for upside-down mounting)
    """
    image = sl.Mat()
    if zed.grab(runtime_params) == sl.ERROR_CODE.SUCCESS:
        zed.retrieve_image(image, sl.VIEW.LEFT)
        # Convert BGRA to BGR
        img_bgra = image.get_data()
        img_bgr = cv2.cvtColor(img_bgra, cv2.COLOR_BGRA2BGR)
        # Rotate if needed (e.g., wrist camera mounted upside down)
        if rotate_180:
            img_bgr = cv2.rotate(img_bgr, cv2.ROTATE_180)
        return img_bgr
    else:
        return None


def bgr_to_rgb(img):
    """Convert BGR to RGB"""
    return cv2.cvtColor(img, cv2.COLOR_BGR2RGB)


def main():
    global _emergency_stop, _zed_ext, _zed_wri
    
    print("=" * 60)
    print("MolmoAct2-Polymetis Bridge (Joint Position)")
    print("=" * 60)
    print(f"Policy: MolmoAct2")
    print(f"NUC IP: {NUC_IP}:4242 (zerorpc)")
    print(f"MolmoAct2 server: {MOLMOACT2_URL}")
    print(f"Task: {PROMPT}")
    print("=" * 60)

    # Verify MolmoAct2 server is reachable
    print("\n[1/4] Connecting to MolmoAct2 server...")
    try:
        resp = requests.get(MOLMOACT2_URL.rstrip("/") + "/act", timeout=5)
        print(f"[OK] MolmoAct2 server reachable (status {resp.status_code})")
    except Exception as e:
        print(f"[WARN] Could not reach MolmoAct2 server: {e}")
        print(f"  Make sure the server is running at {MOLMOACT2_URL}")
        print("  Continuing anyway...")
    
    # Connect to robot (same as GELLO: launch=True will auto-start robot on NUC)
    print("\n[2/4] Connecting to robot on NUC...")
    try:
        from robot.real.server_interface import ServerInterface
        
        # Connect to zerorpc and auto-launch robot (same as GELLO)
        print(f"  Connecting to zerorpc at {NUC_IP}:4242...")
        print("  (This will auto-launch Polymetis robot_server on NUC)")
        robot = ServerInterface(ip_address=NUC_IP)  # launch=True by default!
        print("  [OK] Robot launched and connected")
        
        # Test connection by getting robot state
        print("  Testing robot state...")
        joint_pos = robot.get_joint_positions()
        print(f"  [OK] Got joint positions: {joint_pos[:3]}... (showing first 3)")
        
        # Reset to home position at startup
        print("\n  Resetting robot to home position...")
        home_joints = np.array([
            0.0, -0.50, 0.0, -2.40, 
            0.0, 1.90, 0.0
        ])
        current_joints = robot.get_joint_positions()
        print(f"    Current: {np.round(current_joints, 3).tolist()}")
        print(f"    Target:  {np.round(home_joints, 3).tolist()}")
        
        # Move to home position
        robot.update_joints(
            command=home_joints.tolist(),
            velocity=False,
            blocking=True
        )
        
        # Open gripper (trying 0.0)
        robot.update_gripper(command=0.0, velocity=False, blocking=True)
        print("  [OK] Robot reset to home position (gripper opened)")
        
        print("[OK] Robot interface ready")
        
    except Exception as e:
        print(f"[ERR] Robot connection failed: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Open cameras
    print(f"\n[3/4] Initializing ZED cameras...")
    print("Camera setup: 2x ZED cameras (official OpenPI configuration)")
    print(f"  - ZED 2 (external/shoulder view): SN {ZED_EXTERNAL_SN}")
    print(f"  - ZED Mini (wrist view): SN {ZED_WRIST_SN}")
    
    # Initialize variables
    global _zed_ext, _zed_wri
    zed_ext = None
    zed_ext_runtime = None
    zed_wri = None
    zed_wri_runtime = None
    
    # Initialize external ZED camera (ZED 2)
    try:
        zed_ext, zed_ext_runtime, ext_model, ext_sn = open_zed_camera(
            serial_number=ZED_EXTERNAL_SN,
            width=WIDTH,
            height=HEIGHT,
            fps=FPS
        )
        _zed_ext = zed_ext  # Store for cleanup
        print(f"[OK] External ZED camera (SN: {ZED_EXTERNAL_SN})")
        print(f"  Model: {ext_model}")
        # Warm up camera
        for _ in range(10):
            get_zed_image(zed_ext, zed_ext_runtime)
    except Exception as e:
        print(f"[ERR] External ZED camera initialization failed: {e}")
        print_zed_device_list()
        print_zed_sn_lookup_help()
        cleanup_cameras()
        return
    
    # Initialize wrist ZED camera (ZED Mini)
    try:
        zed_wri, zed_wri_runtime, wri_model, wri_sn = open_zed_camera(
            serial_number=ZED_WRIST_SN,
            width=WIDTH,
            height=HEIGHT,
            fps=FPS
        )
        _zed_wri = zed_wri  # Store for cleanup
        print(f"[OK] Wrist ZED camera (SN: {ZED_WRIST_SN})")
        print(f"  Model: {wri_model}")
        # Warm up camera
        for _ in range(10):
            get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
    except Exception as e:
        print(f"[ERR] Wrist ZED camera initialization failed: {e}")
        print_zed_device_list()
        print_zed_sn_lookup_help()
        cleanup_cameras()
        return
    
    # Preview cameras and wait for confirmation
    print("\n[4/5] Camera Preview")
    print("=" * 60)
    print("Displaying camera feeds...")
    print("Check the camera window to verify:")
    print("  - External ZED 2 (left) shows the scene correctly")
    print("  - Wrist ZED Mini (right) shows the robot gripper")
    print("  - Robot is at home position")
    print("  - Scene is set up correctly")
    print("=" * 60)
    print("\n[READY]  Press ENTER when ready to start inference, or Ctrl+C to quit")
    
    try:
        # Preview loop - show cameras until user presses Enter
        preview_counter = 0
        while True:
            # Collect active camera images
            cameras_to_show = []
            camera_labels = []
            
            # Get External ZED image
            ext_img = None
            if zed_ext:
                ext_img = get_zed_image(zed_ext, zed_ext_runtime)
                if ext_img is not None:
                    cameras_to_show.append(ext_img)
                    camera_labels.append("External (ZED 2)")
            
            # Get Wrist ZED image
            wri_img = None
            if zed_wri:
                wri_img = get_zed_image(zed_wri, zed_wri_runtime)  # No rotation
                if wri_img is not None:
                    cameras_to_show.append(wri_img)
                    camera_labels.append("Wrist (ZED Mini)")
            
            # Skip if no wrist image available
            if wri_img is None or ext_img is None:
                continue
            
            # Only update display every 3 frames for smooth preview
            if preview_counter % 3 == 0 and len(cameras_to_show) > 0:
                # Resize all cameras (doubled size: 848x480 instead of 424x240)
                displays = [cv2.resize(img, (848, 480)) for img in cameras_to_show]
                
                # Stack horizontally
                combined = np.hstack(displays)
                
                # Add text labels (larger font for bigger window)
                for i, label in enumerate(camera_labels):
                    x_pos = 20 + i * 860
                    cv2.putText(combined, label, (x_pos, 50), 
                               cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 255, 0), 3)
                
                # Add instruction text (centered based on number of cameras)
                text_x = 280 if len(cameras_to_show) == 2 else 600
                cv2.putText(combined, "Press ENTER to start", (text_x, 450), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 255), 2)
                
                cv2.imshow("Camera Preview - Press ENTER to start", combined)
            
            preview_counter += 1
            
            # Check for Enter key or window close (check every frame for responsiveness)
            key = cv2.waitKey(1)  # 1ms instead of 30ms for better responsiveness
            if key == 13:  # Enter key
                print("\n[OK] Starting inference...")
                cv2.destroyAllWindows()  # Close the preview window
                break
            elif key == 27:  # ESC key
                print("\n[ERR] Cancelled by user")
                cv2.destroyAllWindows()
                cleanup_cameras()
                return
                
    except KeyboardInterrupt:
        print("\n[ERR] Cancelled by user")
        cv2.destroyAllWindows()
        cleanup_cameras()
        return
    
    dt = 1.0 / CTRL_HZ
    print(f"\nControl Frequency: {CTRL_HZ} Hz")
    print(f"Policy Query Frequency: ~{CTRL_HZ / OPEN_LOOP_HORIZON:.1f} Hz (every {OPEN_LOOP_HORIZON} steps)")
    print("\n[KEYS]  Controls:")
    print("  - Press 'Q' to stop current episode early (saves and continues to next)")
    print("  - Press Ctrl+C to emergency stop (saves and exits program)")
    print("-" * 60)
    
    # ===== Recording Setup =====
    vid_base_dir = Path(__file__).parent / "vid"
    vid_base_dir.mkdir(exist_ok=True)
    
    # Create task-based folder with position for all recording modes
    if LOOP > 0:
        # Convert task to folder name
        task_folder_name = PROMPT.lower().replace(" ", "_").replace(".", "")
        # Structure: vid/task_name/pos-Y/model_name/ (if ENABLE_POSITION_VARIANT) or vid/task_name/model_name/
        if ENABLE_POSITION_VARIANT:
            pos_dir = vid_base_dir / task_folder_name / POSITION_VARIANT
            model_dir = pos_dir / "MolmoAct2"
        else:
            model_dir = vid_base_dir / task_folder_name / "MolmoAct2"
        task_dir = model_dir  # Final directory for this session
        task_dir.mkdir(parents=True, exist_ok=True)
        if LOOP > 1:
            print(f"\n[LOOP] Recording {LOOP} episodes")
        else:
            print(f"\n[LOOP] Recording 1 episode")
        if ENABLE_POSITION_VARIANT:
            print(f"   Position: {POSITION_VARIANT}")
        print(f"   Save to: {task_dir}")
    else:
        print(f"\n[INF]  Infinite inference mode (LOOP=0)")
        print(f"   Press Ctrl+C to stop anytime")
        print(f"   No video recording")
    
    # Determine number of iterations
    infinite_mode = (LOOP == 0)
    num_iterations = float('inf') if infinite_mode else LOOP
    
    # ===== Main recording loop (iterate for each episode) =====
    completed_episodes = 0
    current_episode = 1
    session_model_name = None  # Store model name for entire session (ask once)
    
    while completed_episodes < num_iterations:
        # Check emergency stop at start of each episode
        if _emergency_stop:
            print("\n[STOP] Emergency stop - Exiting episode loop", flush=True)
            break
        
        if LOOP > 1:
            print(f"\n{'='*70}")
            print(f"📹 EPISODE {current_episode} (Completed: {completed_episodes}/{LOOP})")
            print(f"{'='*70}")
        
        step = 0
        last_print_time = time.time()
        
        # Episode-specific preview image (will be captured from first frame)
        episode_preview_image = None
        
        # Open loop control: cache model predictions
        action_queue = []
        
        # EMA smoothing: track smoothed target position
        q_target_smoothed = None
        
        # ===== Setup video recording for this episode (skip if LOOP=0) =====
        if not infinite_mode:
            # Create episode folder with timestamp
            session_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_dir = task_dir / session_timestamp
            session_dir.mkdir(exist_ok=True)
            
            # Save prompt to text file
            instruction_file = session_dir / "instruction.txt"
            with open(instruction_file, "w") as f:
                f.write(f"Session: {session_timestamp}\n")
                f.write(f"Instruction: {PROMPT}\n")
                if LOOP >= 1:
                    f.write(f"Episode: {current_episode} (Progress: {completed_episodes}/{LOOP} completed)\n")
                f.write(f"Steps: TBD\n")
                f.write(f"Control Frequency: {CTRL_HZ} Hz\n")
                f.write(f"Success: TBD\n")  # Will be updated at the end
            
            print(f"\n📹 Video Recording Setup:")
            if LOOP >= 1:
                print(f"   Episode: {current_episode} (Progress: {completed_episodes}/{LOOP} completed)")
            print(f"   Session: {session_timestamp}")
            print(f"   Directory: {session_dir}")
            print(f"   Instruction saved to: {instruction_file}")
            
            # Initialize video writers with mp4v
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            print(f"   Using codec: mp4v (MPEG-4)")
            
            shoulder_video = cv2.VideoWriter(
                str(session_dir / "shoulder_view.mp4"),
                fourcc, FPS, (WIDTH, HEIGHT)  # Use camera FPS (15), not control frequency
            )
            wrist_video = cv2.VideoWriter(
                str(session_dir / "wrist_view.mp4"),
                fourcc, FPS, (WIDTH, HEIGHT)  # Use camera FPS (15), not control frequency
            )
            
            # Verify writers
            if not shoulder_video.isOpened() or not wrist_video.isOpened():
                print(f"   [WARN] ERROR: Failed to open video writers!")
            
            print(f"   Shoulder video: {session_dir / 'shoulder_view.mp4'}")
            print(f"   Wrist video: {session_dir / 'wrist_view.mp4'}")
            print(f"   Recording will stop after {MAX_STEPS} steps\n")
        else:
            print(f"\n[START] Starting inference (no recording)")
            print(f"   Will run until Ctrl+C or max {MAX_STEPS} steps\n")
        
        print(f"\n[5/5] Starting control loop")
        print(f"Frequency: {CTRL_HZ} Hz")
        print(f"Open loop horizon: {OPEN_LOOP_HORIZON} steps (query model every {OPEN_LOOP_HORIZON} steps)")
        print("Press Ctrl+C to stop")
        print("-" * 60)
    
        try:
            while step < MAX_STEPS:
                # Check emergency stop
                if _emergency_stop:
                    print("\n[STOP] Emergency stop detected - Exiting control loop gracefully", flush=True)
                    break
                
                t0 = time.time()
                
                # Check for 'q' key press to stop early (non-blocking)
                import select
                import sys
                if select.select([sys.stdin], [], [], 0.0)[0]:
                    key = sys.stdin.read(1)
                    if key.lower() == 'q':
                        print("\n\n[STOP]  'Q' key pressed - Stopping current episode early")
                        break
                
                # Get images from ZED cameras every step (for smooth video recording)
                ext_img = get_zed_image(zed_ext, zed_ext_runtime)
                wri_img = get_zed_image(zed_wri, zed_wri_runtime)
                
                if ext_img is None or wri_img is None:
                    continue
                
                # Capture first frame as preview for this episode
                if step == 0 and episode_preview_image is None:
                    episode_preview_image = ext_img.copy()
                    print(f"   📸 Captured preview image for episode {current_episode}")
                
                # Write frames to video every step (skip if infinite mode)
                if not infinite_mode:
                    shoulder_video.write(ext_img)
                    wrist_video.write(wri_img)
                
                # Query model when action queue is empty
                if len(action_queue) == 0:
                    
                    # Get current robot state
                    joint_pos = robot.get_joint_positions()
                    gripper_pos = robot.get_gripper_position()
                    
                    # Prepare images: convert BGR→RGB, resize to model input size
                    ext_rgb = bgr_to_rgb(cv2.resize(ext_img, (INPUT_WIDTH, INPUT_HEIGHT)))
                    wri_rgb = bgr_to_rgb(cv2.resize(wri_img, (INPUT_WIDTH, INPUT_HEIGHT)))

                    # Build state: [q1..q7, gripper] shape (8,)
                    state = np.concatenate([joint_pos, [gripper_pos]])

                    # Send request to MolmoAct2 server
                    payload = {
                        "external_cam": ext_rgb,
                        "wrist_cam": wri_rgb,
                        "timestamp": time.time(),
                        "instruction": PROMPT,
                        "state": state,
                    }
                    try:
                        serialized = json_numpy.dumps(payload)
                        response = requests.post(
                            MOLMOACT2_URL.rstrip("/") + "/act",
                            headers={"Content-Type": "application/json"},
                            data=serialized,
                        )
                        if response.status_code != 200:
                            raise RuntimeError(f"Server error {response.status_code}: {response.text}")
                        out = response.json()
                    except Exception as e:
                        if time.time() - last_print_time > 1.0:
                            print(f"[WARN] MolmoAct2 inference error: {e}")
                            last_print_time = time.time()
                        time.sleep(dt)
                        continue

                    # Parse actions
                    if "actions" not in out:
                        print(f"[ERR] No 'actions' key in output: {list(out.keys())}")
                        break

                    actions = np.array(out["actions"])
                    if actions.ndim == 1:
                        actions = actions.reshape(-1, 8)
                    if actions.shape[1] != 8:
                        print(f"[ERR] Unexpected action shape: {actions.shape}, expected (N, 8)")
                        break
                    
                    # Use first OPEN_LOOP_HORIZON actions
                    horizon = min(OPEN_LOOP_HORIZON, len(actions))
                    action_queue = list(actions[:horizon])
                    
                    # if step % 10 == 0:
                    #     print(f"  📥 Queried model: got {len(actions)} actions, using first {horizon}")
                
                # Execute next action from queue
                try:
                    action = action_queue.pop(0)  # Get and remove first action
                    q_model_output = action[:7]  # Joint output from model (either delta or absolute)
                    gripper = action[7]  # gripper command
                    
                    # Get current joint positions
                    q_current = robot.get_joint_positions()
                    
                    # ===== Process Model Output (Delta vs Absolute) =====
                    if OUTPUT_IS_DELTA:
                        # PI0: Model outputs position DELTA (incremental change)
                        dq_raw = q_model_output * DELTA_SCALE  # Scale down delta for safety
                        q_target_raw = q_current + dq_raw  # Calculate target position
                        
                        # No EMA smoothing for deltas - use them directly
                        dq = dq_raw
                        q_target_smoothed = q_target_raw
                    else:
                        # PI05: Model outputs ABSOLUTE position
                        q_target_raw = q_model_output  # Output IS the target position
                        
                        if USE_EMA_SMOOTHING:
                            # Apply EMA smoothing to absolute positions
                            if q_target_smoothed is None:
                                q_target_smoothed = q_target_raw.copy()
                            else:
                                q_target_smoothed = EMA_ALPHA * q_target_raw + (1 - EMA_ALPHA) * q_target_smoothed
                        else:
                            q_target_smoothed = q_target_raw
                        
                        # Calculate delta from smoothed target
                        dq_raw = q_target_raw - q_current
                        dq = q_target_smoothed - q_current
                    
                    # Safety check: clip large movements
                    dq_clipped = np.clip(dq, -MAX_DQ, MAX_DQ)
                    
                    # Use clipped delta
                    q_target = q_current + dq_clipped
                    
                    # Debug: show position and delta
                    # if step % 10 == 0 or step < 5:  # Always show first 5 steps
                    #     if OUTPUT_IS_DELTA:
                    #         print(f"  Current: {q_current[:3].round(3)}... Delta: {dq_raw[:3].round(3)}... Target: {q_target[:3].round(3)}...")
                    #     else:
                    #         print(f"  Current: {q_current[:3].round(3)}... Raw: {q_target_raw[:3].round(3)}... Smoothed: {q_target_smoothed[:3].round(3)}...")
                    #     print(f"  Delta magnitude: {np.abs(dq).max():.3f}, clipped: {np.abs(dq_clipped).max():.3f}")
                    #     print(f"  Actions remaining in queue: {len(action_queue)}")
                    
                    # Send joint command (non-blocking for faster response)
                    robot.update_joints(
                        command=q_target.tolist(),
                        velocity=False,
                        blocking=False
                    )
                    
                    # Gripper: pass model action[7] through unchanged (no clip / smooth / invert)
                    robot.update_gripper(
                        command=float(gripper), velocity=False, blocking=False
                    )
                    
                    # Print progress
                    # if step % 10 == 0 or step < 5:  # Always show first 5 steps
                    #     grip_display = f"{float(gripper):.3f}"
                    #     
                    #     try:
                    #         ee_pose = robot.get_ee_pose()
                    #         ee_pos = ee_pose[:3]
                    #         print(f"Step {step:4d}: δq=[{dq[0]:+.3f},{dq[1]:+.3f},{dq[2]:+.3f},{dq[3]:+.3f},"
                    #               f"{dq[4]:+.3f},{dq[5]:+.3f},{dq[6]:+.3f}] "
                    #               f"EE=[{ee_pos[0]:.3f},{ee_pos[1]:.3f},{ee_pos[2]:.3f}] "
                    #               f"grip={grip_display}")
                    #     except:
                    #         print(f"Step {step:4d}: δq=[{dq[0]:+.3f},{dq[1]:+.3f},{dq[2]:+.3f},{dq[3]:+.3f},"
                    #               f"{dq[4]:+.3f},{dq[5]:+.3f},{dq[6]:+.3f}] grip={grip_display}")
                
                except Exception as e:
                    print(f"[ERR] Action execution error: {e}")
                    import traceback
                    traceback.print_exc()
                    break
                
                step += 1
                
                # Progress bar display
                progress = step / MAX_STEPS * 100
                bar_length = 40
                filled = int(bar_length * step / MAX_STEPS)
                bar = '█' * filled + '-' * (bar_length - filled)
                print(f'\rStep [{bar}] {step}/{MAX_STEPS} ({progress:.1f}%)', end='', flush=True)
                
                elapsed = time.time() - t0
                sleep_time = max(0, dt - elapsed)
                time.sleep(sleep_time)
            
            # Monitor loop time
            if elapsed > dt * 1.5 and step % 10 == 0:
                print(f"[WARN] Loop time too long: {elapsed*1000:.1f}ms (target: {dt*1000:.1f}ms)")
        
        except KeyboardInterrupt:
            print("\n\n[STOP] EMERGENCY STOP - Ctrl+C pressed")
            print("   Stopping robot and saving current episode...")
            
            # Stop robot immediately
            try:
                robot.update_joints(
                    command=robot.get_joint_positions().tolist(),
                    velocity=False,
                    blocking=False
                )
                print("   [OK] Robot stopped")
            except Exception as e:
                print(f"   [WARN] Stop error: {e}")
            
            # Clear action queue
            action_queue.clear()
            
            # Save videos if recording
            if not infinite_mode:
                try:
                    shoulder_video.release()
                    wrist_video.release()
                    print(f"   [OK] Videos saved to: {session_dir}")
                    
                    # Update instruction file (only Steps, Success will be set later)
                    with open(instruction_file, "r") as f:
                        content = f.read()
                    content = content.replace("Steps: TBD", f"Steps: {step}")
                    content = content.replace("Success: TBD", "Success: No")  # Emergency stop = No
                    with open(instruction_file, "w") as f:
                        f.write(content)
                except Exception as e:
                    print(f"   [WARN] Video save error: {e}")
            
            # Return to home
            try:
                print("\n   Moving robot to home position...")
                home_joints = np.array([0.0, -0.5, 0.0, -2.40, 0.0, 1.90, 0.0])
                robot.update_joints(command=home_joints.tolist(), velocity=False, blocking=True)
                # Open gripper (trying 0.0)
                robot.update_gripper(command=0.0, velocity=False, blocking=True)
                print("   [OK] Robot returned to home position (gripper opened)")
            except Exception as e:
                print(f"   [WARN] Reset error: {e}")
            
            # Exit immediately
            cleanup_cameras()
            print("\n[OK] Program terminated by user")
            sys.exit(0)
        
        except Exception as e:
            print(f"\n[ERR] Error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # Close video writers
            if not infinite_mode:
                print("\n📹 Episode Complete!")
                print(f"   Total steps: {step}")
                print(f"   [OK] Videos saved to: {session_dir}")
                shoulder_video.release()
                wrist_video.release()
                
                # Update instruction file with final step count (only if not already updated)
                with open(instruction_file, "r") as f:
                    content = f.read()
                
                modified = False
                # Only update Steps if still TBD (avoid overwriting user's Success input)
                if "Steps: TBD" in content:
                    content = content.replace("Steps: TBD", f"Steps: {step}")
                    modified = True
                # Only set Success to No if still TBD (don't overwrite user input)
                if "Success: TBD" in content:
                    content = content.replace("Success: TBD", "Success: No")
                    modified = True
                
                # Only write if we actually modified something
                if modified:
                    with open(instruction_file, "w") as f:
                        f.write(content)
                    print(f"   [OK] Instruction file updated")
            
            # Reset robot to home position
            print("\n[RESET] Resetting robot...")
            print("  Moving to home position...")
            try:
                home_joints = np.array([0.0, -0.50, 0.0, -2.40, 0.0, 1.90, 0.0])
                current_joints = robot.get_joint_positions()
                print(f"    Current: {np.round(current_joints, 3).tolist()}")
                print(f"    Target:  {np.round(home_joints, 3).tolist()}")
                robot.update_joints(command=home_joints.tolist(), velocity=False, blocking=True)
                # Open gripper (trying 0.0)
                robot.update_gripper(command=0.0, velocity=False, blocking=True)
                print("  [OK] Robot reset complete (gripper opened)")
            except Exception as e:
                print(f"  [ERR] Reset failed: {e}")
        
        # Check for emergency stop before asking for input
        if _emergency_stop:
            print("\n[STOP] Emergency stop detected - Exiting without prompting")
            break
        
        # In infinite mode, just exit after reset
        if infinite_mode:
            print("\n[OK] Episode complete (infinite mode)")
            print("   Exiting program...")
            break
        
        # Ask user what to do next (only for LOOP > 0)
        print(f"\n[PAUSE]  Episode {current_episode} complete (Progress: {completed_episodes}/{LOOP}). Options:")
        if completed_episodes < num_iterations - 1:
            print(f"   Press ENTER to continue to next episode")
            print(f"   Press 'R' + ENTER to re-record this episode")
            print(f"   Press Ctrl+C to exit program")
        else:
            print(f"   Press ENTER to finish (this will complete {LOOP}/{LOOP})")
            print(f"   Press 'R' + ENTER to re-record this episode")
            print(f"   Press Ctrl+C to exit program")
        
        try:
            user_input = input("   Your choice: ").strip().lower()
        except (KeyboardInterrupt, EOFError):
            print("\n\n[STOP] User interrupted - Exiting...")
            _emergency_stop = True
            break
        
        if user_input == 'r':
            print(f"\n[RESET] Re-recording episode {current_episode}...")
            import shutil
            import time as time_module
            time_module.sleep(0.5)
            try:
                if 'session_dir' in locals():
                    print(f"   Deleting: {session_dir}")
                    if session_dir.exists():
                        shutil.rmtree(session_dir)
                        print(f"   [OK] Deleted previous recording")
            except Exception as e:
                print(f"   [ERR] Could not delete: {e}")
        else:
            # User pressed Enter - accept episode
            
            # Excel logging (if enabled)
            if ENABLE_EXCEL_LOGGING and 'session_dir' in locals():
                if not PANDAS_AVAILABLE:
                    print("\n[WARN] Excel logging enabled but pandas not installed. Skipping log.")
                    print("   Install with: pip install pandas openpyxl")
                else:
                    try:
                        # Set model name for logging
                        if session_model_name is None:
                            session_model_name = "MolmoAct2"
                            print(f"\n📝 Model: {session_model_name}")
                        
                        # Ask for success status with retry
                        success_str = None
                        while success_str is None:
                            success_input = input("   Success? (y/n): ").strip().lower()
                            
                            if success_input == 'y':
                                success_str = 'Yes'
                            elif success_input == 'n':
                                success_str = 'No'
                            else:
                                print("   ⊘ Invalid input. Please enter 'y' or 'n'")
                        
                        # Prepare log data (Preview will be added as first column)
                        log_data = {
                            'Task': PROMPT,
                            'Model': session_model_name,
                            'Position': POSITION_VARIANT if ENABLE_POSITION_VARIANT else None,
                            'Episode': current_episode,
                            'Success': success_str,
                            'Video_Path': str((session_dir / 'shoulder_view.mp4').relative_to(Path(__file__).parent)),
                            'Steps': step
                        }
                        
                        # Update instruction.txt with Success value
                        try:
                            with open(instruction_file, "r") as f:
                                content = f.read()
                            # Replace both TBD and No (in case finally block already wrote "No")
                            content = content.replace("Success: TBD", f"Success: {success_str}")
                            content = content.replace("Success: No", f"Success: {success_str}")
                            with open(instruction_file, "w") as f:
                                f.write(content)
                        except:
                            pass
                        
                        # Excel file path in vid directory
                        excel_path = vid_base_dir / 'episode_log.xlsx'
                        
                        # Save to Excel with embedded preview images
                        try:
                            from openpyxl import load_workbook
                            from openpyxl.drawing.image import Image as XLImage
                            from openpyxl import Workbook
                            import io
                            
                            # Define expected column order based on ENABLE_POSITION_VARIANT
                            if ENABLE_POSITION_VARIANT:
                                expected_header = ['Preview', 'Task', 'Model', 'Position', 'Episode', 'Success', 'Video_Path', 'Steps']
                            else:
                                expected_header = ['Preview', 'Task', 'Model', 'Episode', 'Success', 'Video_Path', 'Steps']
                            
                            if excel_path.exists():
                                # Load existing workbook and verify format
                                wb = load_workbook(excel_path)
                                ws = wb.active
                                
                                # Check existing header
                                existing_header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
                                
                                if existing_header != expected_header:
                                    print(f"\n❌ ERROR: Excel format mismatch!")
                                    print(f"   Expected: {expected_header}")
                                    print(f"   Found:    {existing_header}")
                                    print(f"\n   ENABLE_POSITION_VARIANT = {ENABLE_POSITION_VARIANT}")
                                    print(f"   Please check your Excel file format or change ENABLE_POSITION_VARIANT setting.")
                                    print(f"   Program will exit to prevent data corruption.")
                                    import sys
                                    cleanup_cameras()
                                    sys.exit(1)
                                
                                # Append new row data (skip Preview column A)
                                if ENABLE_POSITION_VARIANT:
                                    new_row = [
                                        "",  # Preview column (will add image later)
                                        log_data['Task'],
                                        log_data['Model'],
                                        log_data['Position'],
                                        log_data['Episode'],
                                        log_data['Success'],
                                        log_data['Video_Path'],
                                        log_data['Steps']
                                    ]
                                else:
                                    new_row = [
                                        "",  # Preview column (will add image later)
                                        log_data['Task'],
                                        log_data['Model'],
                                        log_data['Episode'],
                                        log_data['Success'],
                                        log_data['Video_Path'],
                                        log_data['Steps']
                                    ]
                                ws.append(new_row)
                                row_num = ws.max_row
                                
                            else:
                                # Create new workbook
                                wb = Workbook()
                                ws = wb.active
                                
                                # Write header
                                ws.append(expected_header)
                                
                                # Write first data row
                                if ENABLE_POSITION_VARIANT:
                                    ws.append([
                                        "",
                                        log_data['Task'],
                                        log_data['Model'],
                                        log_data['Position'],
                                        log_data['Episode'],
                                        log_data['Success'],
                                        log_data['Video_Path'],
                                        log_data['Steps']
                                    ])
                                else:
                                    ws.append([
                                        "",
                                        log_data['Task'],
                                        log_data['Model'],
                                        log_data['Episode'],
                                        log_data['Success'],
                                        log_data['Video_Path'],
                                        log_data['Steps']
                                    ])
                                row_num = 2  # First data row
                                
                                # Set column width for preview
                                ws.column_dimensions['A'].width = 30
                            
                            # Add preview image for current episode
                            if 'episode_preview_image' in locals() and episode_preview_image is not None:
                                # Resize to preview size
                                preview_img = cv2.resize(episode_preview_image, (320, 180))
                                # Convert BGR to RGB
                                preview_img_rgb = cv2.cvtColor(preview_img, cv2.COLOR_BGR2RGB)
                                # Save to memory buffer
                                from PIL import Image as PILImage
                                pil_img = PILImage.fromarray(preview_img_rgb)
                                img_buffer = io.BytesIO()
                                pil_img.save(img_buffer, format='PNG')
                                img_buffer.seek(0)
                                
                                # Add image to current row
                                ws.row_dimensions[row_num].height = 135
                                img = XLImage(img_buffer)
                                img.width = 320
                                img.height = 180
                                img.anchor = f'A{row_num}'
                                ws.add_image(img)
                            
                            wb.save(excel_path)
                            print(f"   [OK] Logged to: {excel_path}")
                        except ImportError:
                            # Fallback: save without images if openpyxl not available
                            import pandas as pd
                            # Add Preview column for consistency
                            fallback_data = {'Preview': '', **log_data}
                            if excel_path.exists():
                                df = pd.read_excel(excel_path)
                                df = pd.concat([df, pd.DataFrame([fallback_data])], ignore_index=True)
                            else:
                                df = pd.DataFrame([fallback_data])
                            df.to_excel(excel_path, index=False)
                            print(f"   [WARN] openpyxl not available - saved without preview image")
                            print(f"   [OK] Logged to: {excel_path}")
                        except Exception as e:
                            print(f"   [WARN] Image embedding failed: {e}")
                            import traceback
                            traceback.print_exc()
                            # Still save the data even if image embedding fails
                            import pandas as pd
                            # Add Preview column for consistency
                            fallback_data = {'Preview': '', **log_data}
                            if excel_path.exists():
                                df = pd.read_excel(excel_path)
                                df = pd.concat([df, pd.DataFrame([fallback_data])], ignore_index=True)
                            else:
                                df = pd.DataFrame([fallback_data])
                            df.to_excel(excel_path, index=False)
                            print(f"   [OK] Logged to: {excel_path} (data only)")
                    except KeyboardInterrupt:
                        print("\n   ⊘ Logging cancelled")
                    except Exception as e:
                        print(f"\n   [ERR] Logging failed: {e}")
            
            completed_episodes += 1
            current_episode += 1
            print(f"   [OK] Episode accepted ({completed_episodes}/{LOOP} completed)")
    
    # All episodes complete - final cleanup
    print("\n" + "="*70)
    
    if _emergency_stop:
        print("[STOP] Exited due to emergency stop")
        print("   Performing cleanup...")
    elif LOOP > 1:
        print(f"🎉 All {LOOP} episodes completed!")
        print(f"   Videos saved in: {task_dir}")
    elif LOOP == 1:
        print("🎉 Recording complete!")
        if 'session_dir' in locals():
            print(f"   Video saved in: {session_dir}")
    else:
        print("🎉 Session complete (no recording in LOOP=0 mode)")
    print("="*70)
    
    # Final cleanup
    print("\n[Final Cleanup]")
    try:
        print("  [1/2] Moving robot to home position...")
        home_joints = np.array([0.0, -0.5, 0.0, -2.40, 0.0, 1.90, 0.0])
        robot.update_joints(command=home_joints.tolist(), velocity=False, blocking=True)
        robot.update_gripper(command=0.0, velocity=False, blocking=True)
        print("  [OK] Robot at home position")
    except Exception as e:
        print(f"  [ERR] Failed to reset robot: {e}")
    
    print("  [2/2] Closing cameras...")
    cv2.destroyAllWindows()
    cleanup_cameras()
    print("  [OK] Cameras closed")
    print("\n[OK] Program ended successfully")


if __name__ == "__main__":
    main()
